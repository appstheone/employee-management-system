import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog
import csv
import os
import json
import sys
import subprocess
import shutil
from datetime import datetime

from attendence import open_attendance_window

# ============================================================
# OPTIONAL EXCEL SUPPORT
# ============================================================

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True

except ImportError:
    OPENPYXL_AVAILABLE = False


# ============================================================
# PATHS
# ============================================================

BASE_DIR = os.path.dirname(
    os.path.abspath(__file__)
)

CSV_FILE = os.path.join(
    BASE_DIR,
    "employee.csv"
)

USERS_FILE = os.path.join(
    BASE_DIR,
    "users.json"
)

BACKUP_FOLDER = os.path.join(
    BASE_DIR,
    "backups"
)

ACTIVITY_LOG_FILE = os.path.join(
    BASE_DIR,
    "activity_log.csv"
)


# ============================================================
# SETTINGS
# ============================================================

MAX_BACKUPS = 20


# ============================================================
# EMPLOYEE HEADERS
# ============================================================

HEADERS = [
    "Name",
    "Role",
    "Domain",
    "Salary",
    "Joining Date",
    "Project Name"
]


# ============================================================
# ACTIVITY LOG HEADERS
# ============================================================

ACTIVITY_HEADERS = [
    "Date & Time",
    "Username",
    "Role",
    "Action",
    "Details"
]


# ============================================================
# ACTIVITY LOG FUNCTION
# ============================================================

def log_activity(
    username,
    role,
    action,
    details=""
):

    try:

        file_exists = os.path.exists(
            ACTIVITY_LOG_FILE
        )

        with open(
            ACTIVITY_LOG_FILE,
            "a",
            newline="",
            encoding="utf-8"
        ) as file:

            writer = csv.DictWriter(
                file,
                fieldnames=ACTIVITY_HEADERS
            )

            if not file_exists:

                writer.writeheader()

            writer.writerow({
                "Date & Time":
                    datetime.now().strftime(
                        "%Y-%m-%d %H:%M:%S"
                    ),

                "Username":
                    str(username),

                "Role":
                    str(role).upper(),

                "Action":
                    str(action),

                "Details":
                    str(details)
            })

        return True

    except Exception:

        return False


# ============================================================
# LOAD ACTIVITY LOG
# ============================================================

def load_activity_logs():

    logs = []

    if not os.path.exists(
        ACTIVITY_LOG_FILE
    ):
        return logs

    try:

        with open(
            ACTIVITY_LOG_FILE,
            "r",
            newline="",
            encoding="utf-8-sig"
        ) as file:

            reader = csv.DictReader(file)

            for row in reader:

                logs.append({
                    "Date & Time":
                        row.get(
                            "Date & Time",
                            ""
                        ),

                    "Username":
                        row.get(
                            "Username",
                            ""
                        ),

                    "Role":
                        row.get(
                            "Role",
                            ""
                        ),

                    "Action":
                        row.get(
                            "Action",
                            ""
                        ),

                    "Details":
                        row.get(
                            "Details",
                            ""
                        )
                })

    except Exception:
        pass

    return logs


# ============================================================
# LOAD EMPLOYEES
# ============================================================

def load_employees():

    employees = []

    if not os.path.exists(CSV_FILE):

        messagebox.showerror(
            "File Error",
            "employee.csv was not found.\n\n"
            + CSV_FILE
        )

        return employees

    try:

        with open(
            CSV_FILE,
            "r",
            newline="",
            encoding="utf-8-sig"
        ) as file:

            reader = csv.DictReader(file)

            for row in reader:

                employee = {}

                for field in HEADERS:

                    employee[field] = str(
                        row.get(field, "")
                    ).strip()

                employees.append(
                    employee
                )

    except Exception as error:

        messagebox.showerror(
            "CSV Error",
            str(error)
        )

    return employees


# ============================================================
# SAVE EMPLOYEES
# ============================================================

def save_employees(employees):

    try:

        with open(
            CSV_FILE,
            "w",
            newline="",
            encoding="utf-8"
        ) as file:

            writer = csv.DictWriter(
                file,
                fieldnames=HEADERS
            )

            writer.writeheader()

            writer.writerows(
                employees
            )

        return True

    except Exception as error:

        messagebox.showerror(
            "Save Error",
            str(error)
        )

        return False


# ============================================================
# LOAD USERS
# ============================================================

def load_users():

    if not os.path.exists(
        USERS_FILE
    ):

        messagebox.showerror(
            "Error",
            "users.json file not found."
        )

        return {}

    try:

        with open(
            USERS_FILE,
            "r",
            encoding="utf-8"
        ) as file:

            return json.load(file)

    except Exception as error:

        messagebox.showerror(
            "User File Error",
            str(error)
        )

        return {}


# ============================================================
# SAVE USERS
# ============================================================

def save_users(users):

    try:

        with open(
            USERS_FILE,
            "w",
            encoding="utf-8"
        ) as file:

            json.dump(
                users,
                file,
                indent=4
            )

        return True

    except Exception as error:

        messagebox.showerror(
            "User Save Error",
            str(error)
        )

        return False


# ============================================================
# VALIDATE EMPLOYEE
# ============================================================

def validate_employee_data(
    data,
    existing_employees=None,
    current_name=None
):

    if existing_employees is None:
        existing_employees = []

    for field in HEADERS:

        if not str(
            data.get(field, "")
        ).strip():

            return False, (
                f"{field} cannot be empty."
            )

    name = data["Name"].strip()

    if len(name) < 2:

        return False, (
            "Employee name must contain "
            "at least 2 characters."
        )

    for employee in existing_employees:

        existing_name = (
            employee["Name"]
            .strip()
            .lower()
        )

        if current_name is not None:

            if existing_name == (
                current_name
                .strip()
                .lower()
            ):

                continue

        if existing_name == name.lower():

            return False, (
                f"An employee named "
                f"'{name}' already exists."
            )

    salary_text = (
        str(data["Salary"])
        .replace(",", "")
        .replace("₹", "")
        .strip()
    )

    try:

        salary = float(
            salary_text
        )

        if salary < 0:

            return False, (
                "Salary cannot be negative."
            )

    except ValueError:

        return False, (
            "Salary must be a valid number."
        )

    joining_date = (
        data["Joining Date"]
        .strip()
    )

    try:

        datetime.strptime(
            joining_date,
            "%Y-%m-%d"
        )

    except ValueError:

        return False, (
            "Joining Date must use "
            "YYYY-MM-DD format."
        )

    return True, "Valid"


# ============================================================
# CLEAN OLD BACKUPS
# ============================================================

def cleanup_old_backups():

    if not os.path.exists(
        BACKUP_FOLDER
    ):
        return

    try:

        backup_files = []

        for filename in os.listdir(
            BACKUP_FOLDER
        ):

            filepath = os.path.join(
                BACKUP_FOLDER,
                filename
            )

            if (
                os.path.isfile(filepath)
                and filename.lower().endswith(
                    ".csv"
                )
            ):

                try:

                    backup_files.append(
                        (
                            os.path.getmtime(
                                filepath
                            ),
                            filepath
                        )
                    )

                except OSError:
                    pass

        backup_files.sort(
            key=lambda item: item[0],
            reverse=True
        )

        for _, filepath in backup_files[
            MAX_BACKUPS:
        ]:

            try:
                os.remove(filepath)

            except OSError:
                pass

    except Exception:
        pass


# ============================================================
# AUTOMATIC BACKUP
# ============================================================

def create_automatic_backup(
    action="change"
):

    if not os.path.exists(
        CSV_FILE
    ):

        return True

    try:

        os.makedirs(
            BACKUP_FOLDER,
            exist_ok=True
        )

        timestamp = datetime.now().strftime(
            "%Y%m%d_%H%M%S_%f"
        )

        backup_file = os.path.join(
            BACKUP_FOLDER,
            f"auto_{action}_{timestamp}.csv"
        )

        shutil.copy2(
            CSV_FILE,
            backup_file
        )

        cleanup_old_backups()

        return True

    except Exception as error:

        messagebox.showerror(
            "Automatic Backup Error",
            "The data was not changed because "
            "the safety backup could not be created.\n\n"
            f"{error}"
        )

        return False


# ============================================================
# MANUAL BACKUP
# ============================================================

def create_manual_backup():

    if not os.path.exists(
        CSV_FILE
    ):

        return False, (
            "employee.csv was not found."
        )

    try:

        os.makedirs(
            BACKUP_FOLDER,
            exist_ok=True
        )

        timestamp = datetime.now().strftime(
            "%Y%m%d_%H%M%S_%f"
        )

        backup_file = os.path.join(
            BACKUP_FOLDER,
            f"manual_{timestamp}.csv"
        )

        shutil.copy2(
            CSV_FILE,
            backup_file
        )

        cleanup_old_backups()

        return True, backup_file

    except Exception as error:

        return False, str(error)


# ============================================================
# MAIN APPLICATION
# ============================================================

class EmployeeManagementSystem:

    def __init__(
        self,
        root,
        username="admin",
        role="ADMIN"
    ):

        self.root = root

        self.username = (
            str(username).strip()
        )

        self.role = (
            str(role).strip().upper()
        )

        self.employees = []

        # ----------------------------------------------------
        # UNDO DELETE DATA
        # ----------------------------------------------------

        self.last_deleted_employee = None

        self.last_deleted_index = None

        # ----------------------------------------------------
        # WINDOW
        # ----------------------------------------------------

        self.root.title(
            "Employee Management System"
        )

        self.root.geometry(
            "1600x900"
        )

        self.root.minsize(
            1200,
            700
        )

        self.root.protocol(
            "WM_DELETE_WINDOW",
            self.exit_application
        )

        # ----------------------------------------------------
        # CREATE UI
        # ----------------------------------------------------

        self.create_interface()

        self.create_keyboard_shortcuts()

        self.load_data()

        # ----------------------------------------------------
        # LOG CURRENT SESSION
        # ----------------------------------------------------

        log_activity(
            self.username,
            self.role,
            "LOGIN",
            "User opened Employee Management System"
        )

    # ========================================================
    # INTERFACE
    # ========================================================

    def create_interface(self):

        title = tk.Label(
            self.root,
            text="EMPLOYEE MANAGEMENT SYSTEM",
            font=("Arial", 26, "bold")
        )

        title.pack(
            pady=(20, 5)
        )

        login_info = tk.Label(
            self.root,
            text=(
                f"Logged in as: "
                f"{self.username} "
                f"({self.role})"
            ),
            font=("Arial", 13)
        )

        login_info.pack(
            pady=(0, 15)
        )

        # ====================================================
        # DASHBOARD CARDS
        # ====================================================

        cards = tk.Frame(
            self.root
        )

        cards.pack(
            fill="x",
            padx=20,
            pady=10
        )

        self.total_label = (
            self.create_card(
                cards,
                "TOTAL EMPLOYEES"
            )
        )

        self.it_label = (
            self.create_card(
                cards,
                "IT EMPLOYEES"
            )
        )

        self.hr_label = (
            self.create_card(
                cards,
                "HR EMPLOYEES"
            )
        )

        self.total_salary_label = (
            self.create_card(
                cards,
                "TOTAL SALARY"
            )
        )

        self.average_salary_label = (
            self.create_card(
                cards,
                "AVERAGE SALARY"
            )
        )

        self.highest_salary_label = (
            self.create_card(
                cards,
                "HIGHEST SALARY"
            )
        )

        self.lowest_salary_label = (
            self.create_card(
                cards,
                "LOWEST SALARY"
            )
        )

        # ====================================================
        # SEARCH
        # ====================================================

        search_frame = tk.Frame(
            self.root
        )

        search_frame.pack(
            fill="x",
            padx=30,
            pady=(15, 5)
        )

        tk.Label(
            search_frame,
            text="Search:",
            font=("Arial", 12, "bold")
        ).pack(
            side="left"
        )

        self.search_entry = tk.Entry(
            search_frame,
            font=("Arial", 12)
        )

        self.search_entry.pack(
            side="left",
            fill="x",
            expand=True,
            padx=10
        )

        self.search_entry.bind(
            "<Return>",
            lambda event:
            self.search_employee()
        )

        tk.Button(
            search_frame,
            text="Search",
            width=12,
            command=self.search_employee
        ).pack(
            side="left",
            padx=5
        )

        tk.Button(
            search_frame,
            text="Clear",
            width=12,
            command=self.clear_search
        ).pack(
            side="left",
            padx=5
        )

        self.search_status = tk.Label(
            self.root,
            text="Showing all employees",
            font=("Arial", 10)
        )

        self.search_status.pack(
            anchor="w",
            padx=30,
            pady=(0, 5)
        )

        # ====================================================
        # TABLE
        # ====================================================

        table_frame = tk.Frame(
            self.root
        )

        table_frame.pack(
            fill="both",
            expand=True,
            padx=25
        )

        self.tree = ttk.Treeview(
            table_frame,
            columns=HEADERS,
            show="headings"
        )

        widths = {
            "Name": 170,
            "Role": 190,
            "Domain": 110,
            "Salary": 130,
            "Joining Date": 140,
            "Project Name": 260
        }

        for field in HEADERS:

            self.tree.heading(
                field,
                text=field,
                command=lambda col=field:
                self.sort_by_column(
                    col,
                    False
                )
            )

            self.tree.column(
                field,
                width=widths[field],
                anchor="center"
            )

        self.tree.bind(
            "<Double-1>",
            lambda event:
            self.view_employee_info()
        )

        self.tree.bind(
            "<Return>",
            lambda event:
            self.view_employee_info()
        )

        scrollbar_y = ttk.Scrollbar(
            table_frame,
            orient="vertical",
            command=self.tree.yview
        )

        scrollbar_x = ttk.Scrollbar(
            table_frame,
            orient="horizontal",
            command=self.tree.xview
        )

        self.tree.configure(
            yscrollcommand=scrollbar_y.set,
            xscrollcommand=scrollbar_x.set
        )

        self.tree.pack(
            side="left",
            fill="both",
            expand=True
        )

        scrollbar_y.pack(
            side="right",
            fill="y"
        )

        scrollbar_x.pack(
            side="bottom",
            fill="x"
        )

        # ====================================================
        # BUTTONS
        # ====================================================

        button_frame = tk.Frame(
            self.root
        )

        button_frame.pack(
            fill="x",
            padx=25,
            pady=15
        )

        tk.Button(
            button_frame,
            text="Refresh",
            width=12,
            command=self.refresh
        ).pack(
            side="left",
            padx=3
        )

        tk.Button(
            button_frame,
            text="View Info",
            width=12,
            command=self.view_employee_info
        ).pack(
            side="left",
            padx=3
        )

        if self.role == "ADMIN":

            tk.Button(
                button_frame,
                text="Add Employee",
                width=13,
                command=self.add_employee
            ).pack(
                side="left",
                padx=3
            )

            tk.Button(
                button_frame,
                text="Update",
                width=12,
                command=self.update_employee
            ).pack(
                side="left",
                padx=3
            )

            tk.Button(
                button_frame,
                text="Delete",
                width=12,
                command=self.delete_employee
            ).pack(
                side="left",
                padx=3
            )

            # ------------------------------------------------
            # UNDO DELETE
            # ------------------------------------------------

            self.undo_button = tk.Button(
                button_frame,
                text="Undo Delete",
                width=13,
                command=self.undo_last_delete,
                state="disabled"
            )

            self.undo_button.pack(
                side="left",
                padx=3
            )

            tk.Button(
                button_frame,
                text="Manage Users",
                width=13,
                command=self.manage_users
            ).pack(
                side="left",
                padx=3
            )

            tk.Button(
                button_frame,
                text="View Chart",
                width=12,
                command=self.view_chart
            ).pack(
                side="left",
                padx=3
            )

            tk.Button(
                button_frame,
                text="Salary Chart",
                width=12,
                command=self.salary_chart
            ).pack(
                side="left",
                padx=3
            )

            tk.Button(
                button_frame,
                text="Backup Data",
                width=13,
                command=self.backup_data
            ).pack(
                side="left",
                padx=3
            )

            tk.Button(
                button_frame,
                text="Restore Backup",
                width=15,
                command=self.restore_backup
            ).pack(
                side="left",
                padx=3
            )

            tk.Button(
                button_frame,
                text="Attendance",
                width=13,
                command=self.open_attendance
            ).pack(
                side="left",
                padx=3
            )

            tk.Button(
                button_frame,
                text="Activity Log",
                width=14,
                command=self.activity_log_window
            ).pack(
                side="left",
                padx=3
            )

            tk.Button(
                button_frame,
                text="Export Report",
                width=14,
                command=self.export_report
            ).pack(
                side="left",
                padx=3
            )

        tk.Button(
            button_frame,
            text="Change Password",
            width=15,
            command=self.change_password
        ).pack(
            side="left",
            padx=3
        )

        tk.Button(
            button_frame,
            text="Logout",
            width=12,
            command=self.logout
        ).pack(
            side="left",
            padx=3
        )

        tk.Button(
            button_frame,
            text="Exit",
            width=12,
            command=self.exit_application
        ).pack(
            side="left",
            padx=3
        )

        # ====================================================
        # STATUS BAR
        # ====================================================

        self.status_bar = tk.Label(
            self.root,
            text="",
            anchor="w",
            font=("Arial", 10)
        )

        self.status_bar.pack(
            fill="x",
            padx=25,
            pady=(0, 8)
        )

    # ========================================================
    # KEYBOARD SHORTCUTS
    # ========================================================

    def create_keyboard_shortcuts(self):

        self.root.bind(
            "<Control-f>",
            self.shortcut_search
        )

        self.root.bind(
            "<F5>",
            self.shortcut_refresh
        )

        self.root.bind(
            "<Escape>",
            self.shortcut_clear
        )

        self.root.bind(
            "<Control-e>",
            self.shortcut_export
        )

        self.root.bind(
            "<Control-l>",
            self.shortcut_logout
        )

        self.root.bind(
            "<Control-n>",
            self.shortcut_add
        )

        self.root.bind(
            "<Delete>",
            self.shortcut_delete
        )

        self.root.bind(
            "<Control-z>",
            self.shortcut_undo_delete
        )

        # Activity Log shortcut
        self.root.bind(
            "<Control-Shift-l>",
            self.shortcut_activity_log
        )

        # Attendance shortcut
        self.root.bind(
            "<Control-Shift-a>",
            self.shortcut_attendance
        )

    # ========================================================
    # SHORTCUT METHODS
    # ========================================================

    def shortcut_search(
        self,
        event=None
    ):

        self.search_entry.focus_set()

        self.search_entry.select_range(
            0,
            tk.END
        )

        return "break"

    def shortcut_refresh(
        self,
        event=None
    ):

        self.refresh()

        return "break"

    def shortcut_clear(
        self,
        event=None
    ):

        self.clear_search()

        return "break"

    def shortcut_export(
        self,
        event=None
    ):

        if self.role == "ADMIN":

            self.export_report()

        return "break"

    def shortcut_logout(
        self,
        event=None
    ):

        self.logout()

        return "break"

    def shortcut_add(
        self,
        event=None
    ):

        if self.role == "ADMIN":

            self.add_employee()

        return "break"

    def shortcut_delete(
        self,
        event=None
    ):

        if self.role == "ADMIN":

            self.delete_employee()

        return "break"

    def shortcut_undo_delete(
        self,
        event=None
    ):

        if self.role == "ADMIN":

            self.undo_last_delete()

        return "break"

    def shortcut_activity_log(
        self,
        event=None
    ):

        if self.role == "ADMIN":

            self.activity_log_window()

        return "break"

    def shortcut_attendance(
        self,
        event=None
    ):

        if self.role == "ADMIN":

            self.open_attendance()
        else:

            messagebox.showwarning(
                "Access Denied",
                "Only administrators can manage attendance."
            )

        return "break"

    # ========================================================
    # CARD
    # ========================================================

    def create_card(
        self,
        parent,
        title
    ):

        frame = tk.Frame(
            parent,
            relief="solid",
            borderwidth=1
        )

        frame.pack(
            side="left",
            fill="both",
            expand=True,
            padx=4
        )

        tk.Label(
            frame,
            text=title,
            font=("Arial", 11, "bold")
        ).pack(
            pady=(10, 3)
        )

        value = tk.Label(
            frame,
            text="0",
            font=("Arial", 15, "bold")
        )

        value.pack(
            pady=(0, 10)
        )

        return value

    # ========================================================
    # LOAD DATA
    # ========================================================

    def load_data(self):

        self.employees = (
            load_employees()
        )

        self.display_employees(
            self.employees
        )

        self.update_dashboard()

    # ========================================================
    # DISPLAY EMPLOYEES
    # ========================================================

    def display_employees(
        self,
        employees
    ):

        for item in self.tree.get_children():

            self.tree.delete(item)

        for employee in employees:

            self.tree.insert(
                "",
                "end",
                values=(
                    employee["Name"],
                    employee["Role"],
                    employee["Domain"],
                    employee["Salary"],
                    employee["Joining Date"],
                    employee["Project Name"]
                )
            )

        if len(employees) == len(
            self.employees
        ):

            self.search_status.config(
                text=(
                    f"Showing all employees "
                    f"({len(employees)})"
                )
            )

        else:

            self.search_status.config(
                text=(
                    f"Showing {len(employees)} "
                    f"matching employee(s)"
                )
            )

    # ========================================================
    # DASHBOARD
    # ========================================================

    def update_dashboard(self):

        total = len(
            self.employees
        )

        it_count = 0
        hr_count = 0
        total_salary = 0
        salary_list = []

        for employee in self.employees:

            domain = (
                employee["Domain"]
                .strip()
                .upper()
            )

            if domain == "IT":

                it_count += 1

            elif domain == "HR":

                hr_count += 1

            try:

                salary = float(
                    employee["Salary"]
                    .replace(",", "")
                    .replace("₹", "")
                    .strip()
                )

                total_salary += salary

                salary_list.append(
                    (
                        salary,
                        employee["Name"]
                    )
                )

            except (
                ValueError,
                TypeError,
                AttributeError
            ):

                pass

        average_salary = (
            total_salary / total
            if total > 0
            else 0
        )

        if salary_list:

            highest_salary, highest_name = max(
                salary_list,
                key=lambda x: x[0]
            )

            lowest_salary, lowest_name = min(
                salary_list,
                key=lambda x: x[0]
            )

            highest_text = (
                f"₹{highest_salary:,.0f}\n"
                f"{highest_name}"
            )

            lowest_text = (
                f"₹{lowest_salary:,.0f}\n"
                f"{lowest_name}"
            )

        else:

            highest_text = "₹0"

            lowest_text = "₹0"

        self.total_label.config(
            text=str(total)
        )

        self.it_label.config(
            text=str(it_count)
        )

        self.hr_label.config(
            text=str(hr_count)
        )

        self.total_salary_label.config(
            text=f"₹{total_salary:,.0f}"
        )

        self.average_salary_label.config(
            text=f"₹{average_salary:,.0f}"
        )

        self.highest_salary_label.config(
            text=highest_text
        )

        self.lowest_salary_label.config(
            text=lowest_text
        )

        current_time = datetime.now().strftime(
            "%d-%m-%Y %I:%M:%S %p"
        )

        self.status_bar.config(
            text=(
                f"User: {self.username} | "
                f"Role: {self.role} | "
                f"Employees: {total} | "
                f"Last Refresh: {current_time}"
            )
        )

    # ========================================================
    # REFRESH
    # ========================================================

    def refresh(self):

        self.load_data()

        self.search_entry.delete(
            0,
            tk.END
        )

        log_activity(
            self.username,
            self.role,
            "REFRESH",
            "Employee data refreshed"
        )

        self.status_bar.config(
            text=(
                f"Data refreshed | "
                f"Employees: "
                f"{len(self.employees)}"
            )
        )

    # ========================================================
    # SEARCH
    # ========================================================

    def search_employee(self):

        search = (
            self.search_entry
            .get()
            .strip()
            .lower()
        )

        if not search:

            self.display_employees(
                self.employees
            )

            return

        results = []

        for employee in self.employees:

            text = " ".join(
                employee.values()
            ).lower()

            if search in text:

                results.append(
                    employee
                )

        self.display_employees(
            results
        )

        log_activity(
            self.username,
            self.role,
            "SEARCH",
            f"Search term: {search}"
        )

        if not results:

            messagebox.showinfo(
                "Search",
                "No matching employee found."
            )

    # ========================================================
    # CLEAR SEARCH
    # ========================================================

    def clear_search(self):

        self.search_entry.delete(
            0,
            tk.END
        )

        self.display_employees(
            self.employees
        )

    # ========================================================
    # SORT
    # ========================================================

    def sort_by_column(
        self,
        column,
        reverse=False
    ):

        data = []

        for item in self.tree.get_children():

            value = self.tree.set(
                item,
                column
            )

            if column == "Salary":

                try:

                    sort_value = float(
                        str(value)
                        .replace(",", "")
                        .replace("₹", "")
                        .strip()
                    )

                except ValueError:

                    sort_value = 0

            elif column == "Joining Date":

                try:

                    sort_value = (
                        datetime.strptime(
                            value,
                            "%Y-%m-%d"
                        )
                    )

                except ValueError:

                    sort_value = datetime.min

            else:

                sort_value = (
                    str(value).lower()
                )

            data.append(
                (
                    sort_value,
                    item
                )
            )

        data.sort(
            reverse=reverse
        )

        for index, (_, item) in enumerate(
            data
        ):

            self.tree.move(
                item,
                "",
                index
            )

        self.tree.heading(
            column,
            command=lambda:
            self.sort_by_column(
                column,
                not reverse
            )
        )

    # ========================================================
    # GET SELECTED EMPLOYEE
    # ========================================================

    def get_selected_employee(self):

        selected = (
            self.tree.selection()
        )

        if not selected:

            messagebox.showwarning(
                "Selection",
                "Please select an employee."
            )

            return None

        return self.tree.item(
            selected[0]
        )["values"]

    # ========================================================
    # VIEW EMPLOYEE INFO
    # ========================================================

    def view_employee_info(self):

        selected = (
            self.get_selected_employee()
        )

        if selected is None:
            return

        employee_name = str(
            selected[0]
        )

        log_activity(
            self.username,
            self.role,
            "VIEW INFO",
            f"Viewed employee: {employee_name}"
        )

        window = tk.Toplevel(
            self.root
        )

        window.title(
            "Employee Information"
        )

        window.geometry(
            "600x520"
        )

        window.transient(
            self.root
        )

        tk.Label(
            window,
            text="EMPLOYEE INFORMATION",
            font=("Arial", 20, "bold")
        ).pack(
            pady=20
        )

        info_frame = tk.Frame(
            window
        )

        info_frame.pack(
            fill="both",
            expand=True,
            padx=35
        )

        for index, field in enumerate(
            HEADERS
        ):

            tk.Label(
                info_frame,
                text=field + ":",
                font=("Arial", 12, "bold"),
                anchor="w"
            ).grid(
                row=index,
                column=0,
                sticky="w",
                padx=10,
                pady=12
            )

            tk.Label(
                info_frame,
                text=str(
                    selected[index]
                ),
                font=("Arial", 12),
                anchor="w"
            ).grid(
                row=index,
                column=1,
                sticky="w",
                padx=10,
                pady=12
            )

        tk.Button(
            window,
            text="Close",
            width=15,
            command=window.destroy
        ).pack(
            pady=20
        )

    # ========================================================
    # ADD EMPLOYEE
    # ========================================================

    def add_employee(self):

        if self.role != "ADMIN":

            messagebox.showerror(
                "Access Denied",
                "Only administrators can add employees."
            )

            return

        window = tk.Toplevel(
            self.root
        )

        window.title(
            "Add Employee"
        )

        window.geometry(
            "600x560"
        )

        window.transient(
            self.root
        )

        entries = {}

        for index, field in enumerate(
            HEADERS
        ):

            tk.Label(
                window,
                text=field,
                font=("Arial", 11, "bold")
            ).grid(
                row=index,
                column=0,
                padx=15,
                pady=10,
                sticky="w"
            )

            entry = tk.Entry(
                window,
                width=40
            )

            entry.grid(
                row=index,
                column=1,
                padx=15,
                pady=10
            )

            entries[field] = entry

        tk.Label(
            window,
            text="Joining Date format: YYYY-MM-DD",
            font=("Arial", 9)
        ).grid(
            row=6,
            column=1,
            sticky="w",
            padx=15
        )

        def save():

            employee = {}

            for field in HEADERS:

                employee[field] = (
                    entries[field]
                    .get()
                    .strip()
                )

            valid, message = (
                validate_employee_data(
                    employee,
                    self.employees
                )
            )

            if not valid:

                messagebox.showwarning(
                    "Invalid Employee Data",
                    message,
                    parent=window
                )

                return

            # BACKUP BEFORE ADD

            if not create_automatic_backup(
                "add"
            ):

                return

            self.employees.append(
                employee
            )

            if save_employees(
                self.employees
            ):

                log_activity(
                    self.username,
                    self.role,
                    "ADD EMPLOYEE",
                    (
                        f"Added employee: "
                        f"{employee['Name']}"
                    )
                )

                self.load_data()

                window.destroy()

                messagebox.showinfo(
                    "Success",
                    "Employee added successfully.\n\n"
                    "Automatic backup created."
                )

        tk.Button(
            window,
            text="Save Employee",
            width=20,
            command=save
        ).grid(
            row=7,
            column=0,
            columnspan=2,
            pady=25
        )

    # ========================================================
    # UPDATE EMPLOYEE
    # ========================================================

    def update_employee(self):

        if self.role != "ADMIN":

            messagebox.showerror(
                "Access Denied",
                "Only administrators can update employees."
            )

            return

        selected = (
            self.get_selected_employee()
        )

        if selected is None:
            return

        employee_name = str(
            selected[0]
        )

        employee = None

        for item in self.employees:

            if item["Name"] == employee_name:

                employee = item

                break

        if employee is None:

            messagebox.showerror(
                "Error",
                "Employee could not be found."
            )

            return

        old_employee = employee.copy()

        window = tk.Toplevel(
            self.root
        )

        window.title(
            "Update Employee"
        )

        window.geometry(
            "600x560"
        )

        window.transient(
            self.root
        )

        entries = {}

        for index, field in enumerate(
            HEADERS
        ):

            tk.Label(
                window,
                text=field,
                font=("Arial", 11, "bold")
            ).grid(
                row=index,
                column=0,
                padx=15,
                pady=10,
                sticky="w"
            )

            entry = tk.Entry(
                window,
                width=40
            )

            entry.insert(
                0,
                employee[field]
            )

            entry.grid(
                row=index,
                column=1,
                padx=15,
                pady=10
            )

            entries[field] = entry

        tk.Label(
            window,
            text="Joining Date format: YYYY-MM-DD",
            font=("Arial", 9)
        ).grid(
            row=6,
            column=1,
            sticky="w",
            padx=15
        )

        def save_changes():

            updated_employee = {}

            for field in HEADERS:

                updated_employee[field] = (
                    entries[field]
                    .get()
                    .strip()
                )

            valid, message = (
                validate_employee_data(
                    updated_employee,
                    self.employees,
                    current_name=employee_name
                )
            )

            if not valid:

                messagebox.showwarning(
                    "Invalid Employee Data",
                    message,
                    parent=window
                )

                return

            # BACKUP BEFORE UPDATE

            if not create_automatic_backup(
                "update"
            ):

                return

            employee.update(
                updated_employee
            )

            if save_employees(
                self.employees
            ):

                log_activity(
                    self.username,
                    self.role,
                    "UPDATE EMPLOYEE",
                    (
                        f"Updated employee: "
                        f"{old_employee['Name']} "
                        f"-> "
                        f"{updated_employee['Name']}"
                    )
                )

                self.load_data()

                window.destroy()

                messagebox.showinfo(
                    "Success",
                    "Employee updated successfully.\n\n"
                    "Automatic backup created."
                )

        tk.Button(
            window,
            text="Save Changes",
            width=20,
            command=save_changes
        ).grid(
            row=7,
            column=0,
            columnspan=2,
            pady=25
        )

    # ========================================================
    # DELETE EMPLOYEE
    # ========================================================

    def delete_employee(self):

        if self.role != "ADMIN":

            messagebox.showerror(
                "Access Denied",
                "Only administrators can delete employees."
            )

            return

        selected = (
            self.get_selected_employee()
        )

        if selected is None:
            return

        employee_name = str(
            selected[0]
        )

        employee_index = None

        employee_to_delete = None

        for index, employee in enumerate(
            self.employees
        ):

            if employee["Name"] == employee_name:

                employee_index = index

                employee_to_delete = (
                    employee.copy()
                )

                break

        if employee_to_delete is None:

            messagebox.showerror(
                "Error",
                "Employee could not be found."
            )

            return

        answer = messagebox.askyesno(
            "Delete Employee",
            f"Are you sure you want to delete:\n\n"
            f"{employee_name}?\n\n"
            "You can use Ctrl+Z to undo "
            "the last deletion."
        )

        if not answer:
            return

        # BACKUP BEFORE DELETE

        if not create_automatic_backup(
            "delete"
        ):

            return

        # ----------------------------------------------------
        # STORE FOR UNDO
        # ----------------------------------------------------

        self.last_deleted_employee = (
            employee_to_delete
        )

        self.last_deleted_index = (
            employee_index
        )

        # ----------------------------------------------------
        # DELETE
        # ----------------------------------------------------

        del self.employees[
            employee_index
        ]

        if save_employees(
            self.employees
        ):

            log_activity(
                self.username,
                self.role,
                "DELETE EMPLOYEE",
                (
                    f"Deleted employee: "
                    f"{employee_name}"
                )
            )

            self.load_data()

            if hasattr(
                self,
                "undo_button"
            ):

                self.undo_button.config(
                    state="normal"
                )

            messagebox.showinfo(
                "Deleted",
                f"{employee_name} deleted successfully.\n\n"
                "Automatic backup created.\n\n"
                "Press Ctrl+Z or click "
                "'Undo Delete' to restore."
            )

    # ========================================================
    # UNDO LAST DELETE
    # ========================================================

    def undo_last_delete(self):

        if self.role != "ADMIN":

            messagebox.showerror(
                "Access Denied",
                "Only administrators can undo deletions."
            )

            return

        if self.last_deleted_employee is None:

            messagebox.showinfo(
                "Undo Delete",
                "There is no deleted employee to restore."
            )

            return

        employee = (
            self.last_deleted_employee.copy()
        )

        employee_name = (
            employee["Name"]
        )

        # ----------------------------------------------------
        # CHECK DUPLICATE
        # ----------------------------------------------------

        for existing in self.employees:

            if (
                existing["Name"]
                .strip()
                .lower()
                ==
                employee_name
                .strip()
                .lower()
            ):

                messagebox.showwarning(
                    "Cannot Restore",
                    f"An employee named "
                    f"'{employee_name}' "
                    "already exists."
                )

                return

        answer = messagebox.askyesno(
            "Undo Delete",
            f"Restore deleted employee?\n\n"
            f"{employee_name}\n\n"
            "A safety backup will be created first."
        )

        if not answer:
            return

        # SAFETY BACKUP

        if not create_automatic_backup(
            "undo_delete"
        ):

            return

        index = self.last_deleted_index

        if index is None:

            index = len(
                self.employees
            )

        if index < 0:

            index = 0

        if index > len(
            self.employees
        ):

            index = len(
                self.employees
            )

        self.employees.insert(
            index,
            employee
        )

        if save_employees(
            self.employees
        ):

            log_activity(
                self.username,
                self.role,
                "UNDO DELETE",
                (
                    f"Restored employee: "
                    f"{employee_name}"
                )
            )

            self.load_data()

            self.last_deleted_employee = None

            self.last_deleted_index = None

            if hasattr(
                self,
                "undo_button"
            ):

                self.undo_button.config(
                    state="disabled"
                )

            messagebox.showinfo(
                "Undo Successful",
                f"'{employee_name}' "
                "has been restored successfully."
            )

        else:

            try:

                self.employees.remove(
                    employee
                )

            except ValueError:

                pass

    # ========================================================
    # DOMAIN CHART
    # ========================================================

    def view_chart(self):

        if self.role != "ADMIN":

            messagebox.showerror(
                "Access Denied",
                "Only administrators can view charts."
            )

            return

        try:

            import matplotlib.pyplot as plt

        except ImportError:

            messagebox.showerror(
                "Error",
                "Matplotlib is not installed.\n\n"
                "Run:\npip install matplotlib"
            )

            return

        domains = {}

        for employee in self.employees:

            domain = (
                employee["Domain"]
                .strip()
            )

            if domain:

                domains[domain] = (
                    domains.get(
                        domain,
                        0
                    ) + 1
                )

        if not domains:

            messagebox.showinfo(
                "Chart",
                "No employee data available."
            )

            return

        log_activity(
            self.username,
            self.role,
            "VIEW CHART",
            "Opened Employees by Domain chart"
        )

        plt.figure(
            figsize=(9, 5)
        )

        plt.bar(
            list(domains.keys()),
            list(domains.values())
        )

        plt.title(
            "Employees by Domain"
        )

        plt.xlabel(
            "Domain"
        )

        plt.ylabel(
            "Number of Employees"
        )

        plt.tight_layout()

        plt.show()

    # ========================================================
    # SALARY CHART
    # ========================================================

    def salary_chart(self):

        if self.role != "ADMIN":

            messagebox.showerror(
                "Access Denied",
                "Only administrators can view salary charts."
            )

            return

        try:

            import matplotlib.pyplot as plt

        except ImportError:

            messagebox.showerror(
                "Error",
                "Matplotlib is not installed.\n\n"
                "Run:\npip install matplotlib"
            )

            return

        names = []

        salaries = []

        for employee in self.employees:

            try:

                salary = float(
                    employee["Salary"]
                    .replace(",", "")
                    .replace("₹", "")
                    .strip()
                )

                names.append(
                    employee["Name"]
                )

                salaries.append(
                    salary
                )

            except (
                ValueError,
                TypeError,
                AttributeError
            ):

                pass

        if not salaries:

            messagebox.showinfo(
                "Salary Chart",
                "No salary data available."
            )

            return

        log_activity(
            self.username,
            self.role,
            "SALARY CHART",
            "Opened Salary Chart"
        )

        plt.figure(
            figsize=(12, 6)
        )

        plt.bar(
            names,
            salaries
        )

        plt.title(
            "Salary by Employee"
        )

        plt.xlabel(
            "Employee"
        )

        plt.ylabel(
            "Salary (₹)"
        )

        plt.xticks(
            rotation=45,
            ha="right"
        )

        plt.tight_layout()

        plt.show()

    # ========================================================
    # ATTENDANCE
    # ========================================================

    def open_attendance(self):

        if self.role != "ADMIN":

            messagebox.showerror(
                "Access Denied",
                "Only administrators can manage attendance."
            )

            return

        log_activity(
            self.username,
            self.role,
            "OPEN ATTENDANCE",
            "Opened Attendance Management"
        )

        open_attendance_window(
            self.root,
            self.employees,
            self.username,
            self.role,
            log_activity
        )

    # ========================================================
    # BACKUP DATA
    # ========================================================

    def backup_data(self):

        if self.role != "ADMIN":

            messagebox.showerror(
                "Access Denied",
                "Only administrators can create backups."
            )

            return

        success, result = (
            create_manual_backup()
        )

        if success:

            log_activity(
                self.username,
                self.role,
                "CREATE BACKUP",
                (
                    "Manual backup created"
                )
            )

            messagebox.showinfo(
                "Backup Successful",
                "Manual backup created successfully.\n\n"
                f"Maximum backups kept: "
                f"{MAX_BACKUPS}"
            )

        else:

            messagebox.showerror(
                "Backup Error",
                result
            )

    # ========================================================
    # BACKUP MANAGER
    # ========================================================

    def restore_backup(self):

        if self.role != "ADMIN":

            messagebox.showerror(
                "Access Denied",
                "Only administrators can manage backups."
            )

            return

        os.makedirs(
            BACKUP_FOLDER,
            exist_ok=True
        )

        window = tk.Toplevel(
            self.root
        )

        window.title(
            "Backup Manager"
        )

        window.geometry(
            "1100x680"
        )

        window.minsize(
            900,
            550
        )

        window.transient(
            self.root
        )

        tk.Label(
            window,
            text="BACKUP MANAGER",
            font=("Arial", 22, "bold")
        ).pack(
            pady=(20, 5)
        )

        tk.Label(
            window,
            text=(
                f"Latest {MAX_BACKUPS} backups "
                "are retained automatically"
            ),
            font=("Arial", 11)
        ).pack(
            pady=(0, 15)
        )

        table_frame = tk.Frame(
            window
        )

        table_frame.pack(
            fill="both",
            expand=True,
            padx=25,
            pady=10
        )

        backup_columns = (
            "File",
            "Type",
            "Date & Time",
            "Employees",
            "Size"
        )

        backup_tree = ttk.Treeview(
            table_frame,
            columns=backup_columns,
            show="headings"
        )

        backup_tree.heading(
            "File",
            text="Backup File"
        )

        backup_tree.heading(
            "Type",
            text="Type"
        )

        backup_tree.heading(
            "Date & Time",
            text="Date & Time"
        )

        backup_tree.heading(
            "Employees",
            text="Employees"
        )

        backup_tree.heading(
            "Size",
            text="Size"
        )

        backup_tree.column(
            "File",
            width=330
        )

        backup_tree.column(
            "Type",
            width=130,
            anchor="center"
        )

        backup_tree.column(
            "Date & Time",
            width=190,
            anchor="center"
        )

        backup_tree.column(
            "Employees",
            width=100,
            anchor="center"
        )

        backup_tree.column(
            "Size",
            width=120,
            anchor="center"
        )

        scrollbar = ttk.Scrollbar(
            table_frame,
            orient="vertical",
            command=backup_tree.yview
        )

        backup_tree.configure(
            yscrollcommand=scrollbar.set
        )

        backup_tree.pack(
            side="left",
            fill="both",
            expand=True
        )

        scrollbar.pack(
            side="right",
            fill="y"
        )

        backup_records = []

        # ====================================================
        # BACKUP TYPE
        # ====================================================

        def get_backup_type(
            filename
        ):

            lower = filename.lower()

            if lower.startswith(
                "auto_add_"
            ):
                return "AUTO - ADD"

            if lower.startswith(
                "auto_update_"
            ):
                return "AUTO - UPDATE"

            if lower.startswith(
                "auto_delete_"
            ):
                return "AUTO - DELETE"

            if lower.startswith(
                "auto_undo_delete_"
            ):
                return "AUTO - UNDO"

            if lower.startswith(
                "before_restore_"
            ):
                return "RESTORE SAFETY"

            if lower.startswith(
                "manual_"
            ):
                return "MANUAL"

            return "BACKUP"

        # ====================================================
        # COUNT EMPLOYEES
        # ====================================================

        def count_backup_employees(
            filepath
        ):

            try:

                with open(
                    filepath,
                    "r",
                    newline="",
                    encoding="utf-8-sig"
                ) as file:

                    reader = csv.DictReader(
                        file
                    )

                    count = 0

                    for row in reader:

                        if any(
                            str(value).strip()
                            for value in row.values()
                        ):

                            count += 1

                    return count

            except Exception:

                return "N/A"

        # ====================================================
        # FORMAT SIZE
        # ====================================================

        def format_size(
            size
        ):

            if size < 1024:

                return f"{size} B"

            if size < 1024 * 1024:

                return (
                    f"{size / 1024:.1f} KB"
                )

            return (
                f"{size / (1024 * 1024):.1f} MB"
            )

        # ====================================================
        # REFRESH BACKUPS
        # ====================================================

        def refresh_backups():

            backup_records.clear()

            for item in backup_tree.get_children():

                backup_tree.delete(
                    item
                )

            try:

                filenames = os.listdir(
                    BACKUP_FOLDER
                )

            except OSError:

                filenames = []

            valid_files = []

            for filename in filenames:

                filepath = os.path.join(
                    BACKUP_FOLDER,
                    filename
                )

                if (
                    os.path.isfile(filepath)
                    and filename.lower().endswith(
                        ".csv"
                    )
                ):

                    valid_files.append(
                        filename
                    )

            valid_files.sort(
                key=lambda filename:
                os.path.getmtime(
                    os.path.join(
                        BACKUP_FOLDER,
                        filename
                    )
                ),
                reverse=True
            )

            for filename in valid_files:

                filepath = os.path.join(
                    BACKUP_FOLDER,
                    filename
                )

                try:

                    modified = (
                        datetime.fromtimestamp(
                            os.path.getmtime(
                                filepath
                            )
                        )
                    )

                    modified_text = (
                        modified.strftime(
                            "%d-%m-%Y %I:%M:%S %p"
                        )
                    )

                    size = os.path.getsize(
                        filepath
                    )

                    employee_count = (
                        count_backup_employees(
                            filepath
                        )
                    )

                    backup_type = (
                        get_backup_type(
                            filename
                        )
                    )

                    record = {
                        "filename":
                            filename,

                        "path":
                            filepath,

                        "type":
                            backup_type,

                        "datetime":
                            modified,

                        "datetime_text":
                            modified_text,

                        "employees":
                            employee_count,

                        "size":
                            size
                    }

                    backup_records.append(
                        record
                    )

                    backup_tree.insert(
                        "",
                        "end",
                        values=(
                            filename,
                            backup_type,
                            modified_text,
                            employee_count,
                            format_size(
                                size
                            )
                        )
                    )

                except (
                    OSError,
                    ValueError
                ):

                    pass

            backup_status.config(
                text=(
                    f"Total backups: "
                    f"{len(backup_records)} / "
                    f"{MAX_BACKUPS}"
                )
            )

        # ====================================================
        # GET SELECTED BACKUP
        # ====================================================

        def get_selected_backup():

            selected = (
                backup_tree.selection()
            )

            if not selected:

                messagebox.showwarning(
                    "Selection",
                    "Please select a backup first.",
                    parent=window
                )

                return None

            values = backup_tree.item(
                selected[0],
                "values"
            )

            if not values:
                return None

            filename = str(
                values[0]
            )

            for record in backup_records:

                if record[
                    "filename"
                ] == filename:

                    return record

            return None

        # ====================================================
        # VIEW BACKUP INFO
        # ====================================================

        def view_backup_info():

            record = (
                get_selected_backup()
            )

            if record is None:
                return

            messagebox.showinfo(
                "Backup Information",
                f"Backup File:\n"
                f"{record['filename']}\n\n"
                f"Type:\n"
                f"{record['type']}\n\n"
                f"Date & Time:\n"
                f"{record['datetime_text']}\n\n"
                f"Employees:\n"
                f"{record['employees']}\n\n"
                f"File Size:\n"
                f"{format_size(record['size'])}\n\n"
                f"Location:\n"
                f"{record['path']}",
                parent=window
            )

        # ====================================================
        # DELETE BACKUP
        # ====================================================

        def delete_backup():

            record = (
                get_selected_backup()
            )

            if record is None:
                return

            confirm = messagebox.askyesno(
                "Delete Backup",
                "Are you sure you want to delete "
                "this backup?\n\n"
                f"{record['filename']}\n\n"
                "This action cannot be undone.",
                parent=window
            )

            if not confirm:
                return

            try:

                os.remove(
                    record["path"]
                )

                log_activity(
                    self.username,
                    self.role,
                    "DELETE BACKUP",
                    (
                        f"Deleted backup: "
                        f"{record['filename']}"
                    )
                )

                refresh_backups()

                messagebox.showinfo(
                    "Deleted",
                    "Backup deleted successfully.",
                    parent=window
                )

            except Exception as error:

                messagebox.showerror(
                    "Delete Error",
                    str(error),
                    parent=window
                )

        # ====================================================
        # RESTORE SELECTED BACKUP
        # ====================================================

        def restore_selected():

            record = (
                get_selected_backup()
            )

            if record is None:
                return

            confirm = messagebox.askyesno(
                "Confirm Restore",
                "This will replace the current "
                "employee.csv with the selected backup.\n\n"
                f"Backup:\n"
                f"{record['filename']}\n\n"
                "A safety backup will be created first.\n\n"
                "Continue?",
                parent=window
            )

            if not confirm:
                return

            try:

                # SAFETY BACKUP

                if os.path.exists(
                    CSV_FILE
                ):

                    timestamp = (
                        datetime.now().strftime(
                            "%Y%m%d_%H%M%S_%f"
                        )
                    )

                    safety_file = os.path.join(
                        BACKUP_FOLDER,
                        f"before_restore_{timestamp}.csv"
                    )

                    shutil.copy2(
                        CSV_FILE,
                        safety_file
                    )

                # RESTORE

                shutil.copy2(
                    record["path"],
                    CSV_FILE
                )

                self.employees = (
                    load_employees()
                )

                self.display_employees(
                    self.employees
                )

                self.update_dashboard()

                self.last_deleted_employee = None

                self.last_deleted_index = None

                if hasattr(
                    self,
                    "undo_button"
                ):

                    self.undo_button.config(
                        state="disabled"
                    )

                cleanup_old_backups()

                log_activity(
                    self.username,
                    self.role,
                    "RESTORE BACKUP",
                    (
                        f"Restored backup: "
                        f"{record['filename']}"
                    )
                )

                refresh_backups()

                messagebox.showinfo(
                    "Restore Successful",
                    "Employee data restored successfully.\n\n"
                    "A safety backup was created first.",
                    parent=window
                )

            except Exception as error:

                messagebox.showerror(
                    "Restore Error",
                    str(error),
                    parent=window
                )

        # ====================================================
        # OPEN BACKUP FOLDER
        # ====================================================

        def open_backup_folder():

            try:

                if sys.platform.startswith(
                    "win"
                ):

                    os.startfile(
                        BACKUP_FOLDER
                    )

                elif sys.platform == "darwin":

                    subprocess.Popen(
                        [
                            "open",
                            BACKUP_FOLDER
                        ]
                    )

                else:

                    subprocess.Popen(
                        [
                            "xdg-open",
                            BACKUP_FOLDER
                        ]
                    )

                log_activity(
                    self.username,
                    self.role,
                    "OPEN BACKUP FOLDER",
                    "Opened backup folder"
                )

            except Exception as error:

                messagebox.showerror(
                    "Folder Error",
                    str(error),
                    parent=window
                )

        # ====================================================
        # CREATE NEW BACKUP
        # ====================================================

        def create_backup_from_manager():

            success, result = (
                create_manual_backup()
            )

            if success:

                log_activity(
                    self.username,
                    self.role,
                    "CREATE BACKUP",
                    "Manual backup created from Backup Manager"
                )

                refresh_backups()

                messagebox.showinfo(
                    "Backup Successful",
                    "Manual backup created successfully.",
                    parent=window
                )

            else:

                messagebox.showerror(
                    "Backup Error",
                    result,
                    parent=window
                )

        # ====================================================
        # STATUS
        # ====================================================

        backup_status = tk.Label(
            window,
            text="",
            anchor="w",
            font=("Arial", 10)
        )

        backup_status.pack(
            fill="x",
            padx=25,
            pady=(0, 5)
        )

        # ====================================================
        # BUTTONS
        # ====================================================

        button_frame = tk.Frame(
            window
        )

        button_frame.pack(
            fill="x",
            padx=25,
            pady=15
        )

        tk.Button(
            button_frame,
            text="View Info",
            width=14,
            command=view_backup_info
        ).pack(
            side="left",
            padx=4
        )

        tk.Button(
            button_frame,
            text="Restore Selected",
            width=18,
            command=restore_selected
        ).pack(
            side="left",
            padx=4
        )

        tk.Button(
            button_frame,
            text="Delete Backup",
            width=16,
            command=delete_backup
        ).pack(
            side="left",
            padx=4
        )

        tk.Button(
            button_frame,
            text="New Backup",
            width=14,
            command=create_backup_from_manager
        ).pack(
            side="left",
            padx=4
        )

        tk.Button(
            button_frame,
            text="Refresh",
            width=12,
            command=refresh_backups
        ).pack(
            side="left",
            padx=4
        )

        tk.Button(
            button_frame,
            text="Open Folder",
            width=14,
            command=open_backup_folder
        ).pack(
            side="left",
            padx=4
        )

        tk.Button(
            button_frame,
            text="Close",
            width=12,
            command=window.destroy
        ).pack(
            side="right",
            padx=4
        )

        # ====================================================
        # BACKUP WINDOW SHORTCUTS
        # ====================================================

        window.bind(
            "<F5>",
            lambda event:
            refresh_backups()
        )

        window.bind(
            "<Delete>",
            lambda event:
            delete_backup()
        )

        window.bind(
            "<Escape>",
            lambda event:
            window.destroy()
        )

        window.bind(
            "<Return>",
            lambda event:
            restore_selected()
        )

        refresh_backups()

    # ========================================================
    # ACTIVITY LOG WINDOW
    # ========================================================

    def activity_log_window(self):

        if self.role != "ADMIN":

            messagebox.showerror(
                "Access Denied",
                "Only administrators can view the activity log."
            )

            return

        log_activity(
            self.username,
            self.role,
            "VIEW ACTIVITY LOG",
            "Opened Activity Log"
        )

        window = tk.Toplevel(
            self.root
        )

        window.title(
            "Activity Log / Audit Trail"
        )

        window.geometry(
            "1250x700"
        )

        window.minsize(
            950,
            550
        )

        window.transient(
            self.root
        )

        # ====================================================
        # TITLE
        # ====================================================

        tk.Label(
            window,
            text="ACTIVITY LOG / AUDIT TRAIL",
            font=("Arial", 22, "bold")
        ).pack(
            pady=(20, 5)
        )

        tk.Label(
            window,
            text=(
                "Tracks important actions performed "
                "in the Employee Management System"
            ),
            font=("Arial", 11)
        ).pack(
            pady=(0, 15)
        )

        # ====================================================
        # SEARCH
        # ====================================================

        search_frame = tk.Frame(
            window
        )

        search_frame.pack(
            fill="x",
            padx=25,
            pady=5
        )

        tk.Label(
            search_frame,
            text="Search Log:",
            font=("Arial", 11, "bold")
        ).pack(
            side="left"
        )

        log_search_entry = tk.Entry(
            search_frame,
            width=50
        )

        log_search_entry.pack(
            side="left",
            padx=10,
            fill="x",
            expand=True
        )

        # ====================================================
        # TABLE
        # ====================================================

        table_frame = tk.Frame(
            window
        )

        table_frame.pack(
            fill="both",
            expand=True,
            padx=25,
            pady=10
        )

        log_columns = (
            "Date & Time",
            "Username",
            "Role",
            "Action",
            "Details"
        )

        log_tree = ttk.Treeview(
            table_frame,
            columns=log_columns,
            show="headings"
        )

        log_tree.heading(
            "Date & Time",
            text="Date & Time"
        )

        log_tree.heading(
            "Username",
            text="Username"
        )

        log_tree.heading(
            "Role",
            text="Role"
        )

        log_tree.heading(
            "Action",
            text="Action"
        )

        log_tree.heading(
            "Details",
            text="Details"
        )

        log_tree.column(
            "Date & Time",
            width=180,
            anchor="center"
        )

        log_tree.column(
            "Username",
            width=150,
            anchor="center"
        )

        log_tree.column(
            "Role",
            width=100,
            anchor="center"
        )

        log_tree.column(
            "Action",
            width=180,
            anchor="center"
        )

        log_tree.column(
            "Details",
            width=500
        )

        scrollbar_y = ttk.Scrollbar(
            table_frame,
            orient="vertical",
            command=log_tree.yview
        )

        scrollbar_x = ttk.Scrollbar(
            table_frame,
            orient="horizontal",
            command=log_tree.xview
        )

        log_tree.configure(
            yscrollcommand=scrollbar_y.set,
            xscrollcommand=scrollbar_x.set
        )

        log_tree.pack(
            side="left",
            fill="both",
            expand=True
        )

        scrollbar_y.pack(
            side="right",
            fill="y"
        )

        scrollbar_x.pack(
            side="bottom",
            fill="x"
        )

        # ====================================================
        # LOG DATA
        # ====================================================

        all_logs = []

        def display_logs(
            logs
        ):

            for item in log_tree.get_children():

                log_tree.delete(
                    item
                )

            # newest first

            for log in reversed(
                logs
            ):

                log_tree.insert(
                    "",
                    "end",
                    values=(
                        log["Date & Time"],
                        log["Username"],
                        log["Role"],
                        log["Action"],
                        log["Details"]
                    )
                )

            log_status.config(
                text=(
                    f"Total log entries: "
                    f"{len(logs)}"
                )
            )

        # ====================================================
        # REFRESH LOGS
        # ====================================================

        def refresh_logs():

            nonlocal all_logs

            all_logs = load_activity_logs()

            display_logs(
                all_logs
            )

        # ====================================================
        # SEARCH LOGS
        # ====================================================

        def search_logs():

            search = (
                log_search_entry
                .get()
                .strip()
                .lower()
            )

            if not search:

                display_logs(
                    all_logs
                )

                return

            results = []

            for log in all_logs:

                combined = " ".join(
                    [
                        log["Date & Time"],
                        log["Username"],
                        log["Role"],
                        log["Action"],
                        log["Details"]
                    ]
                ).lower()

                if search in combined:

                    results.append(
                        log
                    )

            display_logs(
                results
            )

        # ====================================================
        # VIEW SELECTED LOG
        # ====================================================

        def view_selected_log():

            selected = (
                log_tree.selection()
            )

            if not selected:

                messagebox.showwarning(
                    "Selection",
                    "Please select an activity entry.",
                    parent=window
                )

                return

            values = log_tree.item(
                selected[0],
                "values"
            )

            if not values:
                return

            messagebox.showinfo(
                "Activity Details",
                f"Date & Time:\n"
                f"{values[0]}\n\n"
                f"Username:\n"
                f"{values[1]}\n\n"
                f"Role:\n"
                f"{values[2]}\n\n"
                f"Action:\n"
                f"{values[3]}\n\n"
                f"Details:\n"
                f"{values[4]}",
                parent=window
            )

        # ====================================================
        # EXPORT ACTIVITY LOG
        # ====================================================

        def export_activity_log():

            if not all_logs:

                messagebox.showinfo(
                    "Export",
                    "There are no activity records to export.",
                    parent=window
                )

                return

            filename = filedialog.asksaveasfilename(
                parent=window,
                title="Export Activity Log",
                defaultextension=".csv",
                filetypes=[
                    (
                        "CSV Files",
                        "*.csv"
                    ),
                    (
                        "All Files",
                        "*.*"
                    )
                ],
                initialfile=(
                    "activity_log_"
                    + datetime.now().strftime(
                        "%Y%m%d_%H%M%S"
                    )
                    + ".csv"
                )
            )

            if not filename:
                return

            try:

                with open(
                    filename,
                    "w",
                    newline="",
                    encoding="utf-8"
                ) as file:

                    writer = csv.DictWriter(
                        file,
                        fieldnames=ACTIVITY_HEADERS
                    )

                    writer.writeheader()

                    writer.writerows(
                        all_logs
                    )

                log_activity(
                    self.username,
                    self.role,
                    "EXPORT ACTIVITY LOG",
                    "Exported Activity Log"
                )

                messagebox.showinfo(
                    "Export Successful",
                    "Activity log exported successfully.",
                    parent=window
                )

            except Exception as error:

                messagebox.showerror(
                    "Export Error",
                    str(error),
                    parent=window
                )

        # ====================================================
        # CLEAR LOG
        # ====================================================

        def clear_activity_log():

            confirm = messagebox.askyesno(
                "Clear Activity Log",
                "Are you sure you want to permanently "
                "delete the activity log?\n\n"
                "This action cannot be undone.",
                parent=window
            )

            if not confirm:
                return

            try:

                if os.path.exists(
                    ACTIVITY_LOG_FILE
                ):

                    os.remove(
                        ACTIVITY_LOG_FILE
                    )

                all_logs.clear()

                display_logs(
                    all_logs
                )

                # Create a new entry after clearing.
                log_activity(
                    self.username,
                    self.role,
                    "CLEAR ACTIVITY LOG",
                    "Activity log was cleared"
                )

                refresh_logs()

                messagebox.showinfo(
                    "Activity Log",
                    "Activity log cleared successfully.",
                    parent=window
                )

            except Exception as error:

                messagebox.showerror(
                    "Error",
                    str(error),
                    parent=window
                )

        # ====================================================
        # STATUS
        # ====================================================

        log_status = tk.Label(
            window,
            text="",
            anchor="w",
            font=("Arial", 10)
        )

        log_status.pack(
            fill="x",
            padx=25,
            pady=(0, 5)
        )

        # ====================================================
        # BUTTONS
        # ====================================================

        button_frame = tk.Frame(
            window
        )

        button_frame.pack(
            fill="x",
            padx=25,
            pady=15
        )

        tk.Button(
            search_frame,
            text="Search",
            width=12,
            command=search_logs
        ).pack(
            side="left",
            padx=5
        )

        tk.Button(
            search_frame,
            text="Clear",
            width=12,
            command=lambda: (
                log_search_entry.delete(
                    0,
                    tk.END
                ),
                display_logs(
                    all_logs
                )
            )
        ).pack(
            side="left",
            padx=5
        )

        tk.Button(
            button_frame,
            text="View Selected",
            width=16,
            command=view_selected_log
        ).pack(
            side="left",
            padx=5
        )

        tk.Button(
            button_frame,
            text="Refresh",
            width=14,
            command=refresh_logs
        ).pack(
            side="left",
            padx=5
        )

        tk.Button(
            button_frame,
            text="Export CSV",
            width=14,
            command=export_activity_log
        ).pack(
            side="left",
            padx=5
        )

        tk.Button(
            button_frame,
            text="Clear Log",
            width=14,
            command=clear_activity_log
        ).pack(
            side="left",
            padx=5
        )

        tk.Button(
            button_frame,
            text="Close",
            width=14,
            command=window.destroy
        ).pack(
            side="right",
            padx=5
        )

        # ====================================================
        # ACTIVITY LOG SHORTCUTS
        # ====================================================

        log_search_entry.bind(
            "<Return>",
            lambda event:
            search_logs()
        )

        window.bind(
            "<F5>",
            lambda event:
            refresh_logs()
        )

        window.bind(
            "<Escape>",
            lambda event:
            window.destroy()
        )

        window.bind(
            "<Return>",
            lambda event:
            view_selected_log()
        )

        refresh_logs()

    # ========================================================
    # EXPORT REPORT
    # ========================================================

    def export_report(self):

        if self.role != "ADMIN":

            messagebox.showerror(
                "Access Denied",
                "Only administrators can export reports."
            )

            return

        if not self.employees:

            messagebox.showinfo(
                "Export Report",
                "No employee data available."
            )

            return

        filename = filedialog.asksaveasfilename(
            parent=self.root,
            title="Export Employee Report",
            defaultextension=".xlsx",
            filetypes=[
                (
                    "Excel Files",
                    "*.xlsx"
                ),
                (
                    "CSV Files",
                    "*.csv"
                ),
                (
                    "All Files",
                    "*.*"
                )
            ],
            initialfile=(
                "employee_report_"
                + datetime.now().strftime(
                    "%Y%m%d_%H%M%S"
                )
                + ".xlsx"
            )
        )

        if not filename:
            return

        try:

            if filename.lower().endswith(
                ".xlsx"
            ):

                if not OPENPYXL_AVAILABLE:

                    messagebox.showerror(
                        "Missing Package",
                        "openpyxl is not installed.\n\n"
                        "Run:\npip install openpyxl"
                    )

                    return

                workbook = Workbook()

                sheet = workbook.active

                sheet.title = "Employees"

                for column_number, header in enumerate(
                    HEADERS,
                    start=1
                ):

                    cell = sheet.cell(
                        row=1,
                        column=column_number,
                        value=header
                    )

                    cell.font = Font(
                        bold=True
                    )

                    cell.alignment = Alignment(
                        horizontal="center"
                    )

                for row_number, employee in enumerate(
                    self.employees,
                    start=2
                ):

                    for column_number, header in enumerate(
                        HEADERS,
                        start=1
                    ):

                        cell = sheet.cell(
                            row=row_number,
                            column=column_number,
                            value=employee.get(
                                header,
                                ""
                            )
                        )

                        cell.alignment = Alignment(
                            horizontal="center"
                        )

                for column_number, header in enumerate(
                    HEADERS,
                    start=1
                ):

                    max_length = len(
                        header
                    )

                    for row in sheet.iter_rows(
                        min_col=column_number,
                        max_col=column_number
                    ):

                        for cell in row:

                            if cell.value is not None:

                                max_length = max(
                                    max_length,
                                    len(
                                        str(
                                            cell.value
                                        )
                                    )
                                )

                    sheet.column_dimensions[
                        get_column_letter(
                            column_number
                        )
                    ].width = min(
                        max_length + 4,
                        40
                    )

                sheet.freeze_panes = "A2"

                workbook.save(
                    filename
                )

                log_activity(
                    self.username,
                    self.role,
                    "EXPORT REPORT",
                    f"Exported employee report: {filename}"
                )

                messagebox.showinfo(
                    "Export Successful",
                    "Excel report created successfully."
                )

            else:

                with open(
                    filename,
                    "w",
                    newline="",
                    encoding="utf-8"
                ) as file:

                    writer = csv.DictWriter(
                        file,
                        fieldnames=HEADERS
                    )

                    writer.writeheader()

                    writer.writerows(
                        self.employees
                    )

                log_activity(
                    self.username,
                    self.role,
                    "EXPORT REPORT",
                    f"Exported employee CSV: {filename}"
                )

                messagebox.showinfo(
                    "Export Successful",
                    "CSV report created successfully."
                )

        except Exception as error:

            messagebox.showerror(
                "Export Error",
                str(error)
            )

    # ========================================================
    # MANAGE USERS
    # ========================================================

    def manage_users(self):

        if self.role != "ADMIN":

            messagebox.showerror(
                "Access Denied",
                "Only administrators can manage users."
            )

            return

        users = load_users()

        window = tk.Toplevel(
            self.root
        )

        window.title(
            "Manage Users"
        )

        window.geometry(
            "700x500"
        )

        window.transient(
            self.root
        )

        tk.Label(
            window,
            text="USER MANAGEMENT",
            font=("Arial", 18, "bold")
        ).pack(
            pady=15
        )

        user_tree = ttk.Treeview(
            window,
            columns=(
                "Username",
                "Role"
            ),
            show="headings"
        )

        user_tree.heading(
            "Username",
            text="Username"
        )

        user_tree.heading(
            "Role",
            text="Role"
        )

        user_tree.column(
            "Username",
            width=280,
            anchor="center"
        )

        user_tree.column(
            "Role",
            width=180,
            anchor="center"
        )

        user_tree.pack(
            fill="both",
            expand=True,
            padx=20,
            pady=10
        )

        def refresh_users():

            for item in user_tree.get_children():

                user_tree.delete(
                    item
                )

            current_users = (
                load_users()
            )

            for username, data in (
                current_users.items()
            ):

                if isinstance(
                    data,
                    dict
                ):

                    role = data.get(
                        "role",
                        "EMPLOYEE"
                    )

                else:

                    role = "EMPLOYEE"

                user_tree.insert(
                    "",
                    "end",
                    values=(
                        username,
                        str(role).upper()
                    )
                )

        refresh_users()

        buttons = tk.Frame(
            window
        )

        buttons.pack(
            pady=15
        )

        def add_user():

            username = simpledialog.askstring(
                "Add User",
                "Username:",
                parent=window
            )

            if not username:
                return

            username = username.strip()

            if username in users:

                messagebox.showwarning(
                    "User Exists",
                    "That username already exists.",
                    parent=window
                )

                return

            password = simpledialog.askstring(
                "Add User",
                "Password:",
                show="*",
                parent=window
            )

            if not password:
                return

            role = simpledialog.askstring(
                "Add User",
                "Role (ADMIN or EMPLOYEE):",
                parent=window
            )

            role = (
                role.strip().upper()
                if role
                else "EMPLOYEE"
            )

            if role not in [
                "ADMIN",
                "EMPLOYEE"
            ]:

                role = "EMPLOYEE"

            users[username] = {
                "password": password,
                "role": role
            }

            if save_users(
                users
            ):

                log_activity(
                    self.username,
                    self.role,
                    "ADD USER",
                    f"Created user: {username} ({role})"
                )

                refresh_users()

                messagebox.showinfo(
                    "Success",
                    "User created successfully.",
                    parent=window
                )

        def delete_user():

            selected = (
                user_tree.selection()
            )

            if not selected:

                messagebox.showwarning(
                    "Selection",
                    "Select a user first.",
                    parent=window
                )

                return

            username = user_tree.item(
                selected[0]
            )["values"][0]

            if username == self.username:

                messagebox.showwarning(
                    "Not Allowed",
                    "You cannot delete your current account.",
                    parent=window
                )

                return

            if not messagebox.askyesno(
                "Delete User",
                f"Delete {username}?",
                parent=window
            ):

                return

            if username in users:

                del users[
                    username
                ]

                if save_users(
                    users
                ):

                    log_activity(
                        self.username,
                        self.role,
                        "DELETE USER",
                        f"Deleted user: {username}"
                    )

                    refresh_users()

                    messagebox.showinfo(
                        "Success",
                        "User deleted successfully.",
                        parent=window
                    )

        tk.Button(
            buttons,
            text="Add User",
            width=16,
            command=add_user
        ).pack(
            side="left",
            padx=10
        )

        tk.Button(
            buttons,
            text="Delete User",
            width=16,
            command=delete_user
        ).pack(
            side="left",
            padx=10
        )

        tk.Button(
            buttons,
            text="Close",
            width=16,
            command=window.destroy
        ).pack(
            side="left",
            padx=10
        )

    # ========================================================
    # CHANGE PASSWORD
    # ========================================================

    def change_password(self):

        users = load_users()

        if self.username not in users:

            messagebox.showerror(
                "Error",
                "Current user was not found."
            )

            return

        user_data = users[
            self.username
        ]

        if isinstance(
            user_data,
            dict
        ):

            current_password = str(
                user_data.get(
                    "password",
                    ""
                )
            )

        else:

            current_password = str(
                user_data
            )

        old_password = (
            simpledialog.askstring(
                "Change Password",
                "Current password:",
                show="*",
                parent=self.root
            )
        )

        if old_password is None:
            return

        if old_password != current_password:

            messagebox.showerror(
                "Error",
                "Current password is incorrect."
            )

            return

        new_password = (
            simpledialog.askstring(
                "Change Password",
                "New password:",
                show="*",
                parent=self.root
            )
        )

        if not new_password:
            return

        confirm_password = (
            simpledialog.askstring(
                "Change Password",
                "Confirm new password:",
                show="*",
                parent=self.root
            )
        )

        if new_password != confirm_password:

            messagebox.showerror(
                "Error",
                "Passwords do not match."
            )

            return

        if isinstance(
            user_data,
            dict
        ):

            user_data[
                "password"
            ] = new_password

        else:

            users[
                self.username
            ] = {
                "password":
                    new_password,

                "role":
                    self.role
            }

        if save_users(
            users
        ):

            log_activity(
                self.username,
                self.role,
                "CHANGE PASSWORD",
                "Password changed successfully"
            )

            messagebox.showinfo(
                "Success",
                "Password changed successfully."
            )

    # ========================================================
    # LOGOUT
    # ========================================================

    def logout(self):

        answer = messagebox.askyesno(
            "Logout",
            "Are you sure you want to logout?"
        )

        if not answer:
            return

        log_activity(
            self.username,
            self.role,
            "LOGOUT",
            "User logged out"
        )

        self.root.destroy()

        login_file = os.path.join(
            BASE_DIR,
            "login.py"
        )

        if os.path.exists(
            login_file
        ):

            subprocess.Popen(
                [
                    sys.executable,
                    login_file
                ],
                cwd=BASE_DIR
            )

        else:

            messagebox.showerror(
                "Login Error",
                "login.py was not found."
            )

    # ========================================================
    # EXIT
    # ========================================================

    def exit_application(self):

        answer = messagebox.askyesno(
            "Exit",
            "Are you sure you want to exit?"
        )

        if answer:

            log_activity(
                self.username,
                self.role,
                "EXIT",
                "Application closed"
            )

            self.root.destroy()


# ============================================================
# RUN GUI
# ============================================================

def run_gui(
    username="admin",
    role="ADMIN"
):

    root = tk.Tk()

    EmployeeManagementSystem(
        root,
        username,
        role
    )

    root.mainloop()


# ============================================================
# START
# ============================================================

if __name__ == "__main__":

    if len(sys.argv) >= 3:

        username = sys.argv[1]

        role = sys.argv[2]

    elif len(sys.argv) == 2:

        username = sys.argv[1]

        role = "EMPLOYEE"

    else:

        username = "admin"

        role = "ADMIN"

    run_gui(
        username,
        role
    )